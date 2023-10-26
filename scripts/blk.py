import pandas as pd
import requests
import os
import webbrowser
import http
import aiohttp
import asyncio
import codecs
import re
from openpyxl import Workbook
from typing import List, Dict
import xml.etree.ElementTree as ET
from datetime import datetime


def xml_bytes_to_workbook(xml_bytes) -> Workbook:
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}

    xml_string = xml_bytes.decode("utf-8")
    root = ET.fromstring(xml_string)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for ws in root.findall("ss:Worksheet", namespaces=ns):
        ws_title = ws.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Name")
        worksheet = workbook.create_sheet(title=ws_title)

        for table in ws.findall("ss:Table", namespaces=ns):
            for row in table.findall("ss:Row", namespaces=ns):
                row_cells = []

                for cell in row.findall("ss:Cell", namespaces=ns):
                    cell_data = cell.find("ss:Data", namespaces=ns)
                    cell_value = cell_data.text if cell_data is not None else ""
                    row_cells.append(cell_value)

                worksheet.append(row_cells)

    return workbook


def blk_get_headers(
    url: str,
    cj: http.cookiejar = None,
) -> dict:
    cookie_str = ""
    if cj:
        webbrowser.open("https://www.ishares.com/us/products/etf-investments")
        cookies = {
            cookie.name: cookie.value
            for cookie in cj
            if "ishares" in cookie.domain or "blackrock" in cookie.domain
        }
        cookie_str = "; ".join([f"{key}={value}" for key, value in cookies.items()])
        os.system("taskkill /im chrome.exe /f")

    headers = {
        "authority": "www.ishares.com",
        "method": "GET",
        "path": url,
        "scheme": "https",
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "en-US,en;q=0.9",
        "Cookie": cookie_str,
        "Dnt": "1",
        "Referer": "https://www.ishares.com/us/products/etf-investments",
        "Sec-Ch-Ua": '"Chromium";v="116", "Not)A;Brand";v="24", "Google Chrome";v="116"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": "Windows",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36",
    }
    if not cj:
        del headers["Cookie"]

    return headers


def blk_get_aladdian_info(tickers: List[str], cj: http.cookiejar = None) -> Dict[str, Dict]:
    url = "https://www.ishares.com/us/product-screener/product-screener-v3.1.jsn?dcrPath=/templatedata/config/product-screener-v3/data/en/us-ishares/ishares-product-screener-backend-config&siteEntryPassthrough=true"
    headers = blk_get_headers(url, cj)
    res = requests.get(url, headers=headers)
    json = res.json()

    dict = {}
    for key in json:
        fund = json[key]
        if fund["localExchangeTicker"] in tickers:
            dict[fund["localExchangeTicker"]] = {
                "aladdian_id": key,
                "product_url": fund["productPageUrl"],
                "fund_name": fund["fundName"]
            }
            tickers.remove(fund["localExchangeTicker"])

    if len(tickers) > 0:
        print(f"tickers not found: {tickers}")
    return dict


def blk_get_fund_data(tickers: List[str], raw_path: str, cj: http.cookiejar = None):
    def get_download_links():
        async def fetch(session: aiohttp.ClientSession, url: str, ticker: str) -> pd.DataFrame:
            try:
                headers = blk_get_headers(url, cj)
                full_file_path = os.path.join(
                    raw_path, f"{ticker}_blk_fund_data.xls"
                )
                async with session.get(url, headers=headers) as response:
                    if response.status == 200:
                        bytes = await response.content.read()
                        # https://en.wikipedia.org/wiki/Byte_order_mark
                        bytes = bytes.replace(codecs.BOM_UTF8, b"")
                        wb = xls_to_xlsx(bytes, full_file_path)
                        df = pd.DataFrame(wb)
                        return df
                        
                    else:
                        raise Exception(f"Bad Status: {response.status}")
            except Exception as e:
                print(e)
                return pd.DataFrame()
            
        def remove_invalid_xml_chars(byte_data, encoding='utf-8') -> bytes:
            try:
                xml_str = byte_data.decode(encoding)
            except UnicodeDecodeError as e:
                raise ValueError(f"Could not decode bytes: {e}")

            # XML 1.0 valid characters:
            xml_10_pattern = (
                u'[^\u0009\u000A\u000D\u0020-\uD7FF\uE000-\uFFFD\U00010000-\U0010FFFF]+'
            )
            valid_xml_str = re.sub(xml_10_pattern, '', xml_str)
            valid_byte_data = valid_xml_str.encode(encoding)
            return valid_byte_data
            
        def xls_to_xlsx(bytes: bytes, path: str = None) -> Workbook:
            ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}

            bytes = remove_invalid_xml_chars(bytes)
            xml_string = bytes.decode("utf-8")
            print(xml_string)
            root = ET.fromstring(xml_string)
        
            workbook = Workbook()
            workbook.remove(workbook.active)

            for ws in root.findall("ss:Worksheet", namespaces=ns):
                try:
                    ws_title = ws.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Name")
                    worksheet = workbook.create_sheet(title=ws_title)

                    for table in ws.findall("ss:Table", namespaces=ns):
                        for row in table.findall("ss:Row", namespaces=ns):
                            row_cells = []

                            for cell in row.findall("ss:Cell", namespaces=ns):
                                cell_data = cell.find("ss:Data", namespaces=ns)
                                cell_value = cell_data.text if cell_data is not None else ""
                                row_cells.append(cell_value)

                            worksheet.append(row_cells)
                except:
                    continue
                        
            if path:
                workbook.save(path)
            
            return workbook

        async def get_promises(session: aiohttp.ClientSession):
            aladdin_info = blk_get_aladdian_info(tickers, cj)
            tasks = []
            for ticker in list(aladdin_info.keys()):
                product_url = aladdin_info[ticker]["product_url"]
                ajax = "1521942788811.ajax"
                fund_name_edited = str(aladdin_info[ticker]["fund_name"]).replace(' ', '-')
                url_queries = f"fileType=xls&fileName={fund_name_edited}_fund&dataType=fund" 
                download_url = f"https://www.ishares.com{product_url}/{ajax}?{url_queries}" 
                
                task = fetch(session, download_url, ticker)
                tasks.append(task)

            return await asyncio.gather(*tasks)

        async def run_fetch_all() -> List[pd.DataFrame]:
            async with aiohttp.ClientSession() as session:
                all_data = await get_promises(session)
                return all_data

        
        dfs = asyncio.run(run_fetch_all())
        return dfs
        
    return get_download_links()


def blk_all_funds_info(raw_path: str, cj: http.cookiejar = None) -> pd.DataFrame:
    def get_aladdian_portfolio_ids() -> List[str]:
        path = "https://www.ishares.com/us/product-screener/product-screener-v3.1.jsn?dcrPath=/templatedata/config/product-screener-v3/data/en/us-ishares/ishares-product-screener-backend-config&siteEntryPassthrough=true"
        headers = blk_get_headers(path, cj)
        res = requests.get(path, headers=headers)
        info_dict = res.json()
        return list(info_dict.keys())

    def get_raw_xls_etf_info() -> bytes:
        path = "https://www.ishares.com/us/product-screener/product-screener-v3.1.jsn?type=excel&siteEntryPassthrough=true&dcrPath=/templatedata/config/product-screener-v3/data/en/us-ishares/ishares-product-screener-excel-config&disclosureContentDcrPath=/templatedata/content/article/data/en/us-ishares/DEFAULT/product-screener-all-disclaimer"
        headers = blk_get_headers(path, cj)
        payload = {
            "productView": "etf",
            "portfolios": "-".join(str(x) for x in get_aladdian_portfolio_ids()),
        }
        res = requests.post(path, data=payload, headers=headers, allow_redirects=True)
        return res.content

    wb = xml_bytes_to_workbook(get_raw_xls_etf_info())
    df = pd.DataFrame(wb["etf"].values)
    df.drop(df.tail(3).index, inplace=True)
    df.drop(df.head(2).index, inplace=True)

    cols = [
        "ticker",
        "name",
        "sedol",
        "isin",
        "cusip",
        "inception",
        "gross_expense_ratio",
        "net_expense_ratio",
        "net_asset",
        "net_asset_of_date",
        "asset_class",
        "sub_asset_class",
        "region",
        "market",
        "location",
        "investment_style",
        "12m_trailing_yield",
        "12m_trailing_yield_as_of",
        "ytd_return_pct",
        "ytd_return_pct_as_of",
        "YTD_nav_quarterly",
        "1y_nav_quarterly",
        "3y_nav_quarterly",
        "5y_nav_quarterly",
        "10y_nav_quarterly",
        "inception_nav_quarterly",
        "asof_nav_quarterly",
        "YTD_return_quarterly",
        "1y_return_quarterly",
        "3y_return_quarterly",
        "5y_return_quarterly",
        "10y_return_quarterly",
        "inception_return_quarterly",
        "asof_return_quarterly",
        "YTD_nav_monthly",
        "1y_nav_monthly",
        "3y_nav_monthly",
        "5y_nav_monthly",
        "10y_nav_monthly",
        "inception_nav_monthly",
        "asof_nav_monthly",
        "YTD_return_monthly",
        "1y_return_monthly",
        "3y_return_monthly",
        "5y_return_monthly",
        "10y_return_monthly",
        "inception_return_monthly",
        "asof_return_monthly",
        "12m_trailing_yield",
        "asof_yield_12m",
        "30d_SEC_yield",
        "unsubsidized_yield",
        "asof_yield_30d",
        "Duration_FIC",
        "option_adjusted_spread",
        "avg_yield_FIC_perc",
        "asof_date_FIC",
        "avg_yield_FIC_rating",
    ]
    df = df.iloc[:, :-6]  # frick esg
    df.columns = cols
    df.drop_duplicates(subset=["ticker"], keep="first")

    curr_date = datetime.today().strftime("%Y-%m-%d")
    path = f"{raw_path}/{curr_date}_blk_fund_info.xlsx"
    df.to_excel(
        path,
        index=False,
        sheet_name=f"{curr_date}_blk_fund_info.xlsx",
    )

    return df


if __name__ == "__main__":
    raw_path = r"C:\Users\chris\ETF_Fund_Flows\data\blackrock"
    # blk_all_funds_info(raw_path)
    # info = blk_get_aladdian_info("TLT")
    # print(info)

    # info = blk_get_aladdian_info(["TLT", "TLTW"])
    # for ticker in list(info.keys()):
    #     aladdian_id, product_url = info[ticker]
    #     print(aladdian_id)
    #     print(product_url)
    #     print(info[ticker])

    df = blk_get_fund_data(['TLT'], raw_path)
    print(df)